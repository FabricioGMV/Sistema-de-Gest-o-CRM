from datetime import date, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from model.clientes_model import db, Cliente
from model.servicos_model import ServicoCliente, TipoServico
from model.historico_model import HistoricoCliente

alertas_blueprint = Blueprint('alertas', __name__)


# ── US07: Serviços próximos do vencimento (próximos 15 dias) ─────────────────

@alertas_blueprint.route('/alertas/a-vencer')
def servicos_a_vencer():
    hoje = date.today()
    limite = hoje + timedelta(days=15)
    servicos = (ServicoCliente.query
                .filter(ServicoCliente.data_vencimento >= hoje,
                        ServicoCliente.data_vencimento <= limite,
                        ServicoCliente.status == 'Ativo')
                .order_by(ServicoCliente.data_vencimento)
                .all())
    return render_template('alertas_a_vencer.html', servicos=servicos, hoje=hoje, limite=limite)


# ── US08: Serviços já vencidos ───────────────────────────────────────────────

@alertas_blueprint.route('/alertas/vencidos')
def servicos_vencidos():
    hoje = date.today()
    servicos = (ServicoCliente.query
                .filter(ServicoCliente.data_vencimento < hoje,
                        ServicoCliente.status == 'Ativo')
                .order_by(ServicoCliente.data_vencimento)
                .all())
    return render_template('alertas_vencidos.html', servicos=servicos, hoje=hoje)


# ── US09: Registrar nota rápida no histórico do cliente ──────────────────────

@alertas_blueprint.route('/historico/<int:cliente_id>/novo', methods=['POST'])
def nova_nota(cliente_id):
    texto = request.form.get('texto', '').strip()
    if texto:
        nota = HistoricoCliente(cliente_id=cliente_id, texto=texto, data=date.today())
        db.session.add(nota)
        db.session.commit()
        flash('Nota registrada com sucesso!', 'success')
    else:
        flash('A nota não pode estar vazia.', 'error')
    return redirect(request.referrer or url_for('cliente.listar_clientes'))


@alertas_blueprint.route('/historico/<int:cliente_id>')
def ver_historico(cliente_id):
    cliente = db.session.get(Cliente, cliente_id)
    notas = (HistoricoCliente.query
             .filter_by(cliente_id=cliente_id)
             .order_by(HistoricoCliente.data.desc(), HistoricoCliente.id.desc())
             .all())
    return render_template('historico_cliente.html', cliente=cliente, notas=notas)


# ── US10: Dashboard ──────────────────────────────────────────────────────────

@alertas_blueprint.route('/dashboard')
def dashboard():
    hoje = date.today()
    limite_15 = hoje + timedelta(days=15)
    ultimos_30 = hoje - timedelta(days=30)

    total_clientes = Cliente.query.count()
    total_servicos_ativos = ServicoCliente.query.filter_by(status='Ativo').count()

    a_vencer = (ServicoCliente.query
                .filter(ServicoCliente.data_vencimento >= hoje,
                        ServicoCliente.data_vencimento <= limite_15,
                        ServicoCliente.status == 'Ativo')
                .order_by(ServicoCliente.data_vencimento)
                .all())

    vencidos = (ServicoCliente.query
                .filter(ServicoCliente.data_vencimento < hoje,
                        ServicoCliente.status == 'Ativo')
                .order_by(ServicoCliente.data_vencimento)
                .all())

    recentes = (ServicoCliente.query
                .filter(ServicoCliente.data_execucao >= ultimos_30,
                        ServicoCliente.data_execucao <= hoje)
                .order_by(ServicoCliente.data_execucao.desc())
                .all())

    return render_template('dashboard.html',
                           hoje=hoje,
                           total_clientes=total_clientes,
                           total_servicos_ativos=total_servicos_ativos,
                           a_vencer=a_vencer,
                           vencidos=vencidos,
                           recentes=recentes)


# ── US11: Exportar planilha do mês ───────────────────────────────────────────

@alertas_blueprint.route('/relatorio/mensal')
def relatorio_mensal():
    hoje = date.today()
    primeiro_dia = hoje.replace(day=1)
    if hoje.month == 12:
        ultimo_dia = hoje.replace(year=hoje.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        ultimo_dia = hoje.replace(month=hoje.month + 1, day=1) - timedelta(days=1)

    executados = (ServicoCliente.query
                  .filter(ServicoCliente.data_execucao >= primeiro_dia,
                          ServicoCliente.data_execucao <= ultimo_dia)
                  .order_by(ServicoCliente.data_execucao)
                  .all())

    a_vencer_mes = (ServicoCliente.query
                    .filter(ServicoCliente.data_vencimento >= hoje,
                            ServicoCliente.data_vencimento <= ultimo_dia,
                            ServicoCliente.status == 'Ativo')
                    .order_by(ServicoCliente.data_vencimento)
                    .all())

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = 'Executados no Mês'
    _estilizar_planilha(ws1, executados,
                        titulo=f'Serviços Executados — {hoje.strftime("%B/%Y")}',
                        modo='executados')

    ws2 = wb.create_sheet('A Vencer no Mês')
    _estilizar_planilha(ws2, a_vencer_mes,
                        titulo=f'Serviços a Vencer — {hoje.strftime("%B/%Y")}',
                        modo='a_vencer')

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nome_arquivo = f'relatorio_crm_{hoje.strftime("%Y_%m")}.xlsx'
    return send_file(output,
                     as_attachment=True,
                     download_name=nome_arquivo,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _estilizar_planilha(ws, servicos, titulo, modo):
    verde_escuro = 'FF166534'
    verde_claro  = 'FFDCFCE7'
    amarelo      = 'FFFEF9C3'
    cinza_cabec  = 'FFF1F5F9'
    branco       = 'FFFFFFFF'
    vermelho     = 'FFFEE2E2'

    borda = Border(
        left=Side(style='thin', color='FFE2E8F0'),
        right=Side(style='thin', color='FFE2E8F0'),
        top=Side(style='thin', color='FFE2E8F0'),
        bottom=Side(style='thin', color='FFE2E8F0'),
    )

    # Título na linha 1 (células mescladas A1:F1)
    ws.merge_cells('A1:F1')
    ws['A1'] = titulo
    ws['A1'].font = Font(name='Calibri', bold=True, size=14, color=verde_escuro)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill('solid', fgColor=verde_claro)
    ws.row_dimensions[1].height = 30

    ws.append([])  # linha em branco (linha 2)

    # Cabeçalhos (linha 3)
    if modo == 'executados':
        cabecalhos = ['#', 'Cliente', 'Tipo de Serviço', 'Data Execução', 'Data Vencimento', 'Status']
    else:
        cabecalhos = ['#', 'Cliente', 'Tipo de Serviço', 'Data Vencimento', 'Dias Restantes', 'Situação']

    ws.append(cabecalhos)
    linha_cabec = ws.max_row
    for col, _ in enumerate(cabecalhos, start=1):
        cell = ws.cell(row=linha_cabec, column=col)
        cell.font = Font(bold=True, color='FF475569', name='Calibri', size=10)
        cell.fill = PatternFill('solid', fgColor=cinza_cabec)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = borda
    ws.row_dimensions[linha_cabec].height = 22

    # Dados
    for i, s in enumerate(servicos, start=1):
        if modo == 'executados':
            row = [i, s.cliente.nome, s.tipo.nome,
                   s.data_execucao.strftime('%d/%m/%Y'),
                   s.data_vencimento.strftime('%d/%m/%Y'),
                   s.status]
        else:
            dias = s.dias_para_vencer
            row = [i, s.cliente.nome, s.tipo.nome,
                   s.data_vencimento.strftime('%d/%m/%Y'),
                   dias,
                   'Crítico' if dias <= 7 else 'A Vencer']

        ws.append(row)
        linha_atual = ws.max_row
        cor_fundo = branco if i % 2 == 0 else 'FFF8FAFC'

        if modo == 'a_vencer':
            dias_val = row[4]
            if dias_val <= 7:
                cor_fundo = vermelho
            elif dias_val <= 15:
                cor_fundo = amarelo

        for col in range(1, len(cabecalhos) + 1):
            cell = ws.cell(row=linha_atual, column=col)
            cell.fill = PatternFill('solid', fgColor=cor_fundo)
            cell.alignment = Alignment(horizontal='center' if col != 2 else 'left', vertical='center')
            cell.border = borda
            cell.font = Font(name='Calibri', size=10)

    # Largura das colunas — usa get_column_letter para evitar MergedCell
    larguras = [5, 30, 25, 16, 16, 12]
    for col, larg in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(col)].width = larg

    # Rodapé
    ws.append([])
    ws.append([f'Total de registros: {len(servicos)}'])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, color='FF94A3B8', size=9)